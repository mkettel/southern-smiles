"""
Historical Data Import Script for Southern Smiles Stats & Conditions

Reads the xlsx file and generates a SQL file that can be run in Supabase SQL Editor
to populate all historical stat_entries and oic_log records.

Usage:
  python3 scripts/import-historical.py

Output:
  supabase/historical_data.sql
"""

import openpyxl
from datetime import datetime, timedelta
import re

XLSX_PATH = "/Users/mattkettelkamp/Downloads/Southern Smiles Stats Conditions Chart.xlsx"
OUTPUT_PATH = "/Users/mattkettelkamp/Desktop/southern-smiles/supabase/historical_data.sql"

# ============================================================
# Mapping from spreadsheet names → seed SQL IDs
# ============================================================

STAT_ID_MAP = {
    "BMO": "b0000001-0000-0000-0000-000000000001",
    "Personalized Outflow": "b0000002-0000-0000-0000-000000000002",
    "% Appointments Kept": "b0000003-0000-0000-0000-000000000003",
    "% Recall Appointments Kept": "b0000004-0000-0000-0000-000000000004",
    "Collections": "b0000005-0000-0000-0000-000000000005",
    "Accounts Receivable": "b0000006-0000-0000-0000-000000000006",
    "Collections/Staff": "b0000007-0000-0000-0000-000000000007",
    "# Consults": "b0000008-0000-0000-0000-000000000008",
    "% Tx Plans Presented/Closed": "b0000009-0000-0000-0000-000000000009",
    "Production": "b0000010-0000-0000-0000-000000000010",
    "New Reaches": "b0000011-0000-0000-0000-000000000011",
    "New Patients": "b0000012-0000-0000-0000-000000000012",
}

GOOD_DIRECTION = {
    "BMO": "up",
    "Personalized Outflow": "up",
    "% Appointments Kept": "up",
    "% Recall Appointments Kept": "up",
    "Collections": "up",
    "Accounts Receivable": "down",
    "Collections/Staff": "up",
    "# Consults": "up",
    "% Tx Plans Presented/Closed": "up",
    "Production": "up",
    "New Reaches": "up",
    "New Patients": "up",
}

# These will be replaced with actual UUIDs after auth users are created.
# Using placeholder UUIDs — the generated SQL has a variable substitution section at the top.
EMPLOYEE_PLACEHOLDER = {
    "Dr. Monzer Shakally": "SHAKALLY_UUID",
    "Odalis": "ODALIS_UUID",
    "Lesley": "LESLEY_UUID",
    "Evelis": "EVELIS_UUID",
}

# Map self-reported condition text → enum value
CONDITION_MAP = {
    "Steep Up (Affluence)": "affluence",
    "Steep up (Affluence)": "affluence",
    "Slight Up (Normal)": "normal",
    "Slight up (Normal)": "normal",
    "Flat or Slightly down (Emergency)": "emergency",
    "significant down (Danger)": "danger",
    "Severe down (Non-Existence)": "non_existence",
}


def parse_date(val):
    """Convert various date formats to YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Handle "2026-01-02 00:00:00" format
    match = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if match:
        return match.group(1)
    return None


def to_monday(date_str):
    """Convert a date string to the Monday of that week."""
    if not date_str:
        return None
    d = datetime.strptime(date_str, "%Y-%m-%d")
    # Monday = 0
    monday = d - timedelta(days=d.weekday())
    return monday.strftime("%Y-%m-%d")


def calculate_condition(current, previous, good_direction):
    """Python version of the conditions engine."""
    if previous is None or previous == 0:
        return "normal", 0.0

    pct = ((current - previous) / abs(previous)) * 100
    effective = -pct if good_direction == "down" else pct

    if effective > 20:
        condition = "affluence"
    elif effective > 0:
        condition = "normal"
    elif effective >= -15:
        condition = "emergency"
    elif effective >= -40:
        condition = "danger"
    else:
        condition = "non_existence"

    return condition, round(pct, 4)


def escape_sql(s):
    """Escape a string for SQL single quotes."""
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # ============================================================
    # 1. Parse Normalized_Stats for all stat entries
    # ============================================================
    ws = wb["Normalized_Stats"]
    raw_entries = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not any(v is not None and v != "" for v in row[:6]):
            continue

        employee = row[1]
        stat_name = row[4]
        value = row[5]
        self_cond_raw = row[6]
        week_ending = parse_date(row[8])

        if not employee or not stat_name or value is None or not week_ending:
            continue

        stat_id = STAT_ID_MAP.get(stat_name)
        if not stat_id:
            print(f"WARNING: Unknown stat '{stat_name}', skipping")
            continue

        profile_var = EMPLOYEE_PLACEHOLDER.get(employee)
        if not profile_var:
            print(f"WARNING: Unknown employee '{employee}', skipping")
            continue

        # Convert week_ending (Friday-based) to Monday-based week_start
        week_start = to_monday(week_ending)

        self_condition = CONDITION_MAP.get(self_cond_raw) if self_cond_raw else None

        raw_entries.append({
            "stat_id": stat_id,
            "profile_var": profile_var,
            "week_start": week_start,
            "value": float(value),
            "self_condition": self_condition,
            "stat_name": stat_name,
        })

    # Sort by stat then week to compute previous values
    raw_entries.sort(key=lambda e: (e["stat_id"], e["week_start"]))

    # Compute previous_value, percent_change, auto_condition
    entries_by_stat = {}
    for entry in raw_entries:
        sid = entry["stat_id"]
        if sid not in entries_by_stat:
            entries_by_stat[sid] = []
        entries_by_stat[sid].append(entry)

    final_entries = []
    for sid, stat_entries in entries_by_stat.items():
        stat_name = stat_entries[0]["stat_name"]
        gd = GOOD_DIRECTION.get(stat_name, "up")

        for i, entry in enumerate(stat_entries):
            prev_val = stat_entries[i - 1]["value"] if i > 0 else None
            auto_cond, pct_change = calculate_condition(entry["value"], prev_val, gd)

            entry["previous_value"] = prev_val
            entry["percent_change"] = pct_change
            entry["auto_condition"] = auto_cond
            final_entries.append(entry)

    # ============================================================
    # 2. Parse playbook responses from raw entry sheets
    # ============================================================
    playbook_responses = {}  # key: (stat_name_prefix, timestamp_str) -> response

    raw_sheet_configs = [
        {
            "sheet": "Raw_Entries_Dr",
            "stats": [
                {"col": 1, "name": "Production"},
            ],
            "playbook_cols": {
                3: "affluence",
                4: "normal",
                5: "emergency",
                6: "danger",
                7: "non_existence",
            },
        },
        {
            "sheet": "Raw_Entires_Financial Coordinat",
            "stats": [
                {"col": 1, "name": "Collections", "pb_offset": 0},
                {"col": 8, "name": "Accounts Receivable", "pb_offset": 7},
                {"col": 15, "name": "Collections/Staff", "pb_offset": 14},
            ],
        },
        {
            "sheet": "Raw_Entries_PR Officer",
            "stats": [
                {"col": 1, "name": "New Reaches", "pb_offset": 0},
                {"col": 8, "name": "New Patients", "pb_offset": 7},
            ],
        },
        {
            "sheet": "Raw_Entries_Receptionist",
            "stats": [
                {"col": 1, "name": "BMO", "pb_offset": 0},
                {"col": 8, "name": "Personalized Outflow", "pb_offset": 7},
            ],
        },
        {
            "sheet": "Raw_Entires_Scheduling Coordina",
            "stats": [
                {"col": 1, "name": "% Appointments Kept", "pb_offset": 0},
                {"col": 8, "name": "% Recall Appointments Kept", "pb_offset": 7},
            ],
        },
        {
            "sheet": "Raw_Entries_TX Coordinator",
            "stats": [
                {"col": 1, "name": "# Consults", "pb_offset": 0},
                {"col": 8, "name": "% Tx Plans Presented/Closed", "pb_offset": 7},
            ],
        },
    ]

    for config in raw_sheet_configs:
        ws = wb[config["sheet"]]
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            timestamp = row[0]
            if not timestamp:
                continue
            ts_str = parse_date(timestamp)
            if not ts_str:
                continue
            week_start = to_monday(ts_str)

            for stat_cfg in config["stats"]:
                stat_name = stat_cfg["name"]
                stat_id = STAT_ID_MAP.get(stat_name)
                if not stat_id:
                    continue

                # Find playbook columns for this stat
                # They follow the pattern: value, condition, affluence, normal, emergency, danger, non_existence
                base_col = stat_cfg["col"]
                pb_conditions = ["affluence", "normal", "emergency", "danger", "non_existence"]

                for j, cond in enumerate(pb_conditions):
                    pb_col = base_col + 2 + j  # skip value col and condition col
                    if pb_col < len(row) and row[pb_col] is not None:
                        response = str(row[pb_col]).strip()
                        if response:
                            key = (stat_id, week_start)
                            if key not in playbook_responses:
                                playbook_responses[key] = response
                            # Keep first non-empty response per stat/week

    # Attach playbook responses to entries
    for entry in final_entries:
        key = (entry["stat_id"], entry["week_start"])
        entry["playbook_response"] = playbook_responses.get(key)

    # ============================================================
    # 3. Parse OIC_Log
    # ============================================================
    ws = wb["OIC_Log"]
    oic_entries = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not any(v is not None and v != "" for v in row):
            continue

        effective_date = parse_date(row[2])
        submitted_by = str(row[3]).strip() if row[3] else None
        area = str(row[4]).strip() if row[4] else None
        post_affected = str(row[5]).strip() if row[5] else None
        description = str(row[6]).strip() if row[6] else None

        if not effective_date or not description:
            continue

        # Map submitted_by to profile variable
        profile_var = None
        if submitted_by:
            name_lower = submitted_by.lower()
            if "monzer" in name_lower or "shakally" in name_lower:
                profile_var = "SHAKALLY_UUID"
            elif "lesley" in name_lower:
                profile_var = "LESLEY_UUID"
            elif "odalis" in name_lower:
                profile_var = "ODALIS_UUID"
            elif "evelis" in name_lower:
                profile_var = "EVELIS_UUID"
            else:
                profile_var = "SHAKALLY_UUID"  # default to admin

        oic_entries.append({
            "profile_var": profile_var,
            "effective_date": effective_date,
            "area": area,
            "post_affected": post_affected,
            "description": description,
        })

    # ============================================================
    # 4. Generate SQL
    # ============================================================
    lines = []
    lines.append("-- ============================================================")
    lines.append("-- Southern Smiles — Historical Data Import")
    lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"-- Stat entries: {len(final_entries)}")
    lines.append(f"-- OIC log entries: {len(oic_entries)}")
    lines.append("-- ============================================================")
    lines.append("")
    lines.append("-- IMPORTANT: Replace these placeholder UUIDs with the actual")
    lines.append("-- auth.users UUIDs after creating the users in Supabase.")
    lines.append("-- You can find them in the Supabase dashboard under Authentication > Users.")
    lines.append("")
    lines.append("-- Step 1: Set these variables to your actual user UUIDs")
    lines.append("DO $$")
    lines.append("DECLARE")
    lines.append("  shakally_id uuid;")
    lines.append("  odalis_id uuid;")
    lines.append("  lesley_id uuid;")
    lines.append("  evelis_id uuid;")
    lines.append("BEGIN")
    lines.append("  -- Look up profile IDs by email (adjust emails as needed)")
    lines.append("  SELECT id INTO shakally_id FROM profiles WHERE email ILIKE '%shakally%' OR email ILIKE '%monzer%' OR role = 'admin' LIMIT 1;")
    lines.append("  SELECT id INTO odalis_id FROM profiles WHERE email ILIKE '%odalis%' LIMIT 1;")
    lines.append("  SELECT id INTO lesley_id FROM profiles WHERE email ILIKE '%lesley%' LIMIT 1;")
    lines.append("  SELECT id INTO evelis_id FROM profiles WHERE email ILIKE '%evelis%' LIMIT 1;")
    lines.append("")
    lines.append("  -- Verify we found all profiles")
    lines.append("  IF shakally_id IS NULL THEN RAISE EXCEPTION 'Could not find Shakally profile. Create auth users first.'; END IF;")
    lines.append("  IF odalis_id IS NULL THEN RAISE EXCEPTION 'Could not find Odalis profile. Create auth users first.'; END IF;")
    lines.append("  IF lesley_id IS NULL THEN RAISE EXCEPTION 'Could not find Lesley profile. Create auth users first.'; END IF;")
    lines.append("  IF evelis_id IS NULL THEN RAISE EXCEPTION 'Could not find Evelis profile. Create auth users first.'; END IF;")
    lines.append("")
    lines.append("  RAISE NOTICE 'Found all profiles: Shakally=%, Odalis=%, Lesley=%, Evelis=%', shakally_id, odalis_id, lesley_id, evelis_id;")
    lines.append("")

    # Employee post assignments
    lines.append("  -- ============================================================")
    lines.append("  -- Employee Post Assignments")
    lines.append("  -- ============================================================")
    assignments = [
        ("evelis_id", "a0000001-0000-0000-0000-000000000001", "Evelis → Receptionist"),
        ("lesley_id", "a0000002-0000-0000-0000-000000000002", "Lesley → Scheduling Coordinator"),
        ("lesley_id", "a0000006-0000-0000-0000-000000000006", "Lesley → PR Officer"),
        ("odalis_id", "a0000003-0000-0000-0000-000000000003", "Odalis → Financial Coordinator"),
        ("odalis_id", "a0000004-0000-0000-0000-000000000004", "Odalis → TX Coordinator"),
        ("shakally_id", "a0000005-0000-0000-0000-000000000005", "Dr. Shakally → Doctor"),
    ]
    for var, post_id, comment in assignments:
        lines.append(f"  -- {comment}")
        lines.append(f"  INSERT INTO employee_posts (profile_id, post_id) VALUES ({var}, '{post_id}') ON CONFLICT (profile_id, post_id) DO NOTHING;")
    lines.append("")

    # Stat entries
    lines.append("  -- ============================================================")
    lines.append(f"  -- Stat Entries ({len(final_entries)} records)")
    lines.append("  -- ============================================================")

    for entry in final_entries:
        profile_var = entry["profile_var"].lower().replace("_uuid", "_id")
        stat_id = entry["stat_id"]
        week_start = entry["week_start"]
        value = entry["value"]
        prev_val = entry["previous_value"]
        pct_change = entry["percent_change"]
        auto_cond = entry["auto_condition"]
        self_cond = entry["self_condition"]
        pb_response = entry.get("playbook_response")

        prev_sql = f"{prev_val}" if prev_val is not None else "NULL"
        self_cond_sql = f"'{self_cond}'" if self_cond else "NULL"
        pb_sql = escape_sql(pb_response) if pb_response else "NULL"

        lines.append(
            f"  INSERT INTO stat_entries (stat_id, profile_id, week_start, value, previous_value, percent_change, auto_condition, self_condition, playbook_response)"
            f" VALUES ('{stat_id}', {profile_var}, '{week_start}', {value}, {prev_sql}, {pct_change}, '{auto_cond}', {self_cond_sql}, {pb_sql})"
            f" ON CONFLICT (stat_id, week_start) DO UPDATE SET value = EXCLUDED.value, previous_value = EXCLUDED.previous_value, percent_change = EXCLUDED.percent_change, auto_condition = EXCLUDED.auto_condition, self_condition = EXCLUDED.self_condition, playbook_response = EXCLUDED.playbook_response;"
        )

    lines.append("")

    # OIC Log entries
    lines.append("  -- ============================================================")
    lines.append(f"  -- OIC Log Entries ({len(oic_entries)} records)")
    lines.append("  -- ============================================================")

    for entry in oic_entries:
        profile_var = entry["profile_var"].lower().replace("_uuid", "_id")
        effective_date = entry["effective_date"]
        area = escape_sql(entry["area"])
        post_affected = escape_sql(entry["post_affected"])
        description = escape_sql(entry["description"])

        lines.append(
            f"  INSERT INTO oic_log (profile_id, effective_date, area, post_affected, entry_text)"
            f" VALUES ({profile_var}, '{effective_date}', {area}, {post_affected}, {description});"
        )

    lines.append("")
    lines.append("  RAISE NOTICE 'Historical import complete!';")
    lines.append("END $$;")

    # Write output
    with open(OUTPUT_PATH, "w") as f:
        f.write("\n".join(lines))

    print(f"\nGenerated: {OUTPUT_PATH}")
    print(f"  - {len(final_entries)} stat entries")
    print(f"  - {len(oic_entries)} OIC log entries")
    print(f"  - {len(playbook_responses)} playbook responses attached")
    print(f"  - {len(assignments)} employee post assignments")
    print()

    # Summary by stat
    from collections import Counter
    stat_counts = Counter(e["stat_name"] for e in final_entries)
    print("Entries per stat:")
    for stat, count in sorted(stat_counts.items()):
        print(f"  {stat}: {count}")

    # Summary by employee
    emp_counts = Counter(e["profile_var"] for e in final_entries)
    print("\nEntries per employee:")
    for emp, count in sorted(emp_counts.items()):
        print(f"  {emp}: {count}")


if __name__ == "__main__":
    main()
